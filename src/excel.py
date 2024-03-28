import openpyxl


def get_product_ids(file_name: str, first_row: bool = False) -> set:
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    ids = set()
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True)):
        if not first_row and not row_idx:
            continue
        ids.add(row[0])

    return ids
